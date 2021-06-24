import base64
from itertools import islice
import tempfile

from django.core.exceptions import PermissionDenied
from django.db import models, transaction
from django.http import HttpResponse
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.utils.text import slugify
from django.utils.translation import gettext as _
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from rest_framework.response import Response
from wagtail.admin import messages
from wagtail.admin.views.pages.utils import get_valid_next_url_from_request
from wagtail_localize.models import Translation, TranslationContext, String, StringTranslation, StringSegment
from wagtail_localize.models import UnknownString, UnknownContext, StringNotUsedInContext
from wagtail_localize.views.edit_translation import user_can_edit_instance

# A lot of the code is duplicated from the corresponding PO functionality,
# see upload_pofile and download_pofile from
# https://github.com/wagtail/wagtail-localize/blob/28e2c9fd4b87e90facbf00c6c234e97f75129b57/wagtail_localize/views/edit_translation.py
# and Translation.import_po / Translation.export_po / TranslationSource.export_po
# https://github.com/wagtail/wagtail-localize/blob/28e2c9fd4b87e90facbf00c6c234e97f75129b57/wagtail_localize/models.py
# Ideally a lot of it should be factored out, but I don't see
# how to do this without refactoring wagtail_localize itself.

# TODO: Should these headers be localized?
META_ENTRIES = ["Page UUID", "Timestamp"]
HEADER_ROW = ["ID", "Original", "Translation"]
HEADER_ROW_INDEX = 3  # 1-based offset
FIRST_COLUMN_WIDTH = 12
CONTENT_COLUMN_WIDTH = 80
DEFAULT_CELL_HEIGHT = 15


def uuid_to_base64(uid):
    return base64.b64encode(uid.bytes, altchars=b'_.').decode()


def translation_to_xlsx(translation):
    """
    Exports all translatable strings with any translations that have already been made.

    Returns:
        openpyxl.WorkBook: A spreadsheet containing
            string identifiers (1st column),
            source translatable strings (2nd column),
            translations (3rd column).
    """

    msgs = []

    string_segments = (
        StringSegment.objects.filter(source=translation.source)
        .order_by('order')
        .select_related("context", "string")
        .annotate_translation(translation.target_locale, include_errors=True)
    )

    for string_segment in string_segments:
        msgs.append((string_segment.string.data, string_segment.context.path, string_segment.translation))

    wb = Workbook()
    ws = wb.active
    # We cannot use the canonical string version of the uuid here as a sheet
    # name as the character limit for sheet names is 31 in MS Excel.
    ws.title = uuid_to_base64(translation.uuid)
    ws['A1'] = META_ENTRIES[0]
    ws['B1'] = str(translation.uuid)
    ws['A2'] = META_ENTRIES[1]
    ws['B2'] = str(timezone.now())

    ws.append(HEADER_ROW)

    for text, context, str_translation in msgs:
        ws.append([context, text, str_translation or ""])

    # Add any obsolete segments that have translations for future reference
    # We find this by looking for obsolete contexts and annotate the latest
    # translation for each one. Contexts that were never translated are
    # excluded
    for string_translation in (
        StringTranslation.objects
        .filter(context__object_id=translation.source.object_id, locale=translation.target_locale)
        .exclude(translation_of_id__in=StringSegment.objects.filter(source=translation.source).values_list('string_id', flat=True))
        .select_related("translation_of", "context")
        .iterator()
    ):
        ws.append([string_translation.context.path, string_translation.translation_of.data, string_translation.data or "", "OBSOLETE"])

    # Add styling to the sheet.
    wb.security.lockStructure = True
    ws.protection.sheet = True
    for row in ws['1:2']:
        for cell in row:
            cell.font = Font(italic=True)
    for cell in ws[str(HEADER_ROW_INDEX)]:
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(border_style='thick', color='000000'))

    for cell in ws['C']:
        if int(cell.row) > HEADER_ROW_INDEX:
            cell.fill = PatternFill('solid', fgColor='FFFF00')
            cell.alignment = Alignment(wrap_text=True)
            cell.protection = Protection(locked=False)
    for cell in ws['B']:
        if int(cell.row) > HEADER_ROW_INDEX:
            cell.alignment = Alignment(wrap_text=True)
    for cell in ws['A']:
        if int(cell.row) > HEADER_ROW_INDEX:
            cell.border = Border(right=Side(border_style='thick', color='000000'))
    ws['A' + str(HEADER_ROW_INDEX)].border = Border(
            right=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000'))
    
    # Determine cell width and height
    ws.column_dimensions['A'].width = FIRST_COLUMN_WIDTH
    ws.column_dimensions['B'].width = CONTENT_COLUMN_WIDTH
    ws.column_dimensions['C'].width = CONTENT_COLUMN_WIDTH
    for row in islice(ws.rows, HEADER_ROW_INDEX, None):
        # Process rows after the header row
        # (islice uses a 0-based index offset)
        length = max(len(str(cell.value or '')) for cell in row)
        height = DEFAULT_CELL_HEIGHT * (length//CONTENT_COLUMN_WIDTH + 1)
        ws.row_dimensions[row[0].row].height = height

    return wb


def download_xlsxfile(request, translation_id):
    translation = get_object_or_404(Translation, id=translation_id)

    instance = translation.get_target_instance()
    if not user_can_edit_instance(request.user, instance):
        raise PermissionDenied

    wb = translation_to_xlsx(translation)
    with tempfile.NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        data = tmp.read()

    response = HttpResponse(data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = (
        "attachment; filename=%s-%s.xlsx" % (
            slugify(translation.source.object_repr),
            translation.target_locale.language_code,
        )
    )
    return response


def xlsx_to_translation(translation, wb, delete=False, user=None, translation_type='manual', tool_name=""):
    """
    Imports all translatable strings with any translations that have already been made.
    Args:
        wb (openpyxl.Workbook): A XLSX Workbook object containing the source translatable strings and any translations.
        delete (boolean, optional): Set to True to delete any translations that do not appear in the XLSX file.
        user (User, optional): The user who is performing this operation. Used for logging purposes.
        translation_type ('manual' or 'machine', optional): Whether the translationw as performed by a human or machine. Defaults to 'manual'.
        tool_name (string, optional): The name of the tool that was used to perform the translation. Defaults to ''.
    Returns:
        list[POImportWarning]: A list of POImportWarning objects representing any non-fatal issues that were
        encountered while importing the XLSX file.
    """

    # TODO: UnknownString, UnknownContext, StringNotUsedInContext
    #       are subclasses of POImportWarning.
    #       I don't see a good way of making analogues of these for XLSX
    #       without code duplication or refactoring of wagtail_localize.
    #       The list of returned warnings doesn't seem to be used currently
    #       anyway, so it might not be important at this point.
    # TODO: The equivalent method for PO files is annotated with @transaction.atomic
    #       Ensure that the writing here is atomic as well?
    #       https://github.com/wagtail/wagtail-localize/blob/28e2c9fd4b87e90facbf00c6c234e97f75129b57/wagtail_localize/models.py#L1041

    seen_translation_ids = set()
    warnings = []
    ws = wb[uuid_to_base64(translation.uuid)]

    sheet_header_row = [c.value for c in ws[str(HEADER_ROW_INDEX)]]
    if ws['B1'].value != str(translation.uuid) or sheet_header_row != HEADER_ROW:
        # TODO: Warn the user that the spreadsheet format seems to be off?
        return []

    for index, row in enumerate(ws.iter_rows(min_row=HEADER_ROW_INDEX+1)):
        msgctxt = row[0].value
        msgid = row[1].value
        msgstr = row[2].value
        try:
            string = String.objects.get(locale_id=translation.source.locale_id, data=msgid)
            context = TranslationContext.objects.get(object_id=translation.source.object_id, path=msgctxt)

            # Ignore blank strings
            if not msgstr:
                continue

            # Ignore if the string doesn't appear in this context, and if there is not an obsolete StringTranslation
            if not StringSegment.objects.filter(string=string, context=context).exists() and not StringTranslation.objects.filter(translation_of=string, context=context).exists():
                warnings.append(StringNotUsedInContext(index, msgid, msgctxt))
                continue

            string_translation, created = string.translations.get_or_create(
                locale_id=translation.target_locale_id,
                context=context,
                defaults={
                    "data": msgstr,
                    "updated_at": timezone.now(),
                    "translation_type": translation_type,
                    "tool_name": tool_name,
                    'last_translated_by': user,
                    'has_error': False,
                    'field_error': "",
                },
            )

            seen_translation_ids.add(string_translation.id)

            if not created:
                # Update the string_translation only if it has changed
                if string_translation.data != msgstr:
                    string_translation.data = msgstr
                    string_translation.translation_type = translation_type
                    string_translation.tool_name = tool_name
                    string_translation.last_translated_by = user
                    string_translation.updated_at = timezone.now()
                    string_translation.save()

        except TranslationContext.DoesNotExist:
            warnings.append(UnknownContext(index, msgctxt))

        except String.DoesNotExist:
            warnings.append(UnknownString(index, msgid))

    # Delete any translations that weren't mentioned
    if delete:
        StringTranslation.objects.filter(context__object_id=translation.source.object_id, locale=translation.target_locale).exclude(id__in=seen_translation_ids).delete()

    return warnings


@csrf_exempt  # TODO: This is a quick hack for testing only
@require_POST
def upload_xlsxfile(request, translation_id):
    translation = get_object_or_404(Translation, id=translation_id)

    instance = translation.get_target_instance()
    if not user_can_edit_instance(request.user, instance):
        raise PermissionDenied

    do_import = True

    with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
        f.write(request.FILES["file"].read())
        f.flush()

        try:
            wb = load_workbook(f.name)
        except:
            # openpoxl doesn't provide a specific exception for invalid
            # input files. It just crashes in whichever step it fails,
            # with a variety of possible errors, e.g.
            # zipfile.BadZipFile: File is not a zip file
            # KeyError: "There is no item named '[Content_Types].xml' in the archive"
            # xml.etree.ElementTree.ParseError: not well-formed (invalid token)
            messages.error(
                request,
                _("Please upload a valid XLSX file.")
            )
            do_import = False

    if do_import:
        # On Windows, it seems to be possible to change the sheet title
        # of the exported worksheet (in LibreOffice you cannot).
        # We might consider looking for the UUID entry in cell A2
        # if no sheet with the right title exists, as a backup.
        if not uuid_to_base64(translation.uuid) in wb.sheetnames:
            messages.error(
                request,
                _("Cannot import XLSX file that was created for a different translation.")
            )
            do_import = False

    if do_import:
        xlsx_to_translation(translation, wb, user=request.user, tool_name="XLSX File")

        messages.success(
            request,
            _("Successfully imported translations from XLSX File.")
        )

    # Work out where to redirect to
    next_url = get_valid_next_url_from_request(request)
    if not next_url:
        # Note: You should always provide a next URL when using this view!
        next_url = reverse('wagtailadmin_home')

    return redirect(next_url)
